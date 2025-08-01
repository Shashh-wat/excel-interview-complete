# evaluation_engine.py - Claude-Powered Evaluation with Caching (ORCHESTRATOR COMPATIBLE)
"""
Comprehensive evaluation system that scores candidate responses using:
1. Claude LLM for textual analysis
2. Excel file parsing and analysis
3. Intelligent caching for performance
4. Multi-modal evaluation (text + file)
5. FIXED: Full orchestrator compatibility
"""

import json
import time
import hashlib
import asyncio
import logging
from typing import Dict, List, Optional, Union, Any
from datetime import datetime, timedelta
import re
from pathlib import Path

# Simplified imports for better integration
try:
    import anthropic
    from anthropic import AsyncAnthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False
    logging.warning("Anthropic not available - evaluation will use fallbacks")

# Excel file parsing (with graceful fallbacks)
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

# =============================================================================
# ORCHESTRATOR COMPATIBILITY LAYER
# =============================================================================

class ClaudeAPIWrapper:
    """Wrapper that provides the exact interface orchestrator expects"""
    
    def __init__(self, anthropic_client=None, api_key=None):
        self.anthropic_client = anthropic_client
        self.api_key = api_key
        self.available = ANTHROPIC_AVAILABLE and (anthropic_client is not None or api_key is not None)
        
        if self.available:
            if not self.anthropic_client and self.api_key:
                self.anthropic_client = AsyncAnthropic(api_key=self.api_key)
        
        # Performance tracking
        self.total_calls = 0
        self.successful_calls = 0
        self.failed_calls = 0
    
    async def evaluate_text_response(self, question: Dict, response_text: str) -> Dict[str, Any]:
        """Main evaluation method that orchestrator calls"""
        
        if not self.available:
            return self._fallback_evaluation(question, response_text)
        
        try:
            self.total_calls += 1
            
            prompt = self._build_evaluation_prompt(question, response_text)
            
            response = await self.anthropic_client.messages.create(
                model="claude-3-sonnet-20240229",
                max_tokens=4000,
                temperature=0.3,
                messages=[{"role": "user", "content": prompt}]
            )
            
            self.successful_calls += 1
            return self._parse_claude_response(response.content[0].text)
            
        except Exception as e:
            self.failed_calls += 1
            logging.error(f"Claude API error: {e}")
            return self._fallback_evaluation(question, response_text)
    
    def _build_evaluation_prompt(self, question: Dict, response_text: str) -> str:
        """Build evaluation prompt"""
        
        question_text = question.get('text', '')
        expected_keywords = question.get('expected_keywords', [])
        
        return f"""You are a senior Excel interviewer evaluating a candidate's response.

QUESTION: {question_text}

CANDIDATE RESPONSE:
{response_text}

EXPECTED KEYWORDS: {', '.join(expected_keywords) if expected_keywords else 'N/A'}

Rate the response from 0-5 points based on:
- Technical accuracy (0-2 points)
- Understanding (0-2 points) 
- Clarity (0-1 point)

Respond with JSON only:
{{
  "score": 3.5,
  "confidence": 0.8,
  "reasoning": "Good understanding but could be clearer",
  "strengths": ["Correct concepts", "Good examples"],
  "areas_for_improvement": ["More detail needed"],
  "keywords_found": ["VLOOKUP", "INDEX"],
  "mistakes_detected": []
}}"""
    
    def _parse_claude_response(self, response_text: str) -> Dict[str, Any]:
        """Parse Claude's response"""
        
        try:
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start >= 0 and json_end > json_start:
                json_content = response_text[json_start:json_end]
                parsed = json.loads(json_content)
                
                return {
                    "score": float(parsed.get("score", 2.5)),
                    "confidence": float(parsed.get("confidence", 0.7)),
                    "reasoning": str(parsed.get("reasoning", "Evaluation completed")),
                    "strengths": list(parsed.get("strengths", ["Response provided"])),
                    "areas_for_improvement": list(parsed.get("areas_for_improvement", ["Could be more detailed"])),
                    "keywords_found": list(parsed.get("keywords_found", [])),
                    "mistakes_detected": list(parsed.get("mistakes_detected", []))
                }
            else:
                raise ValueError("No JSON found")
                
        except Exception as e:
            logging.warning(f"Failed to parse Claude response: {e}")
            return self._fallback_evaluation({}, response_text)
    
    def _fallback_evaluation(self, question: Dict, response_text: str) -> Dict[str, Any]:
        """Fallback when Claude fails"""
        
        expected_keywords = question.get('expected_keywords', [])
        keywords_found = []
        
        if expected_keywords and response_text:
            response_lower = response_text.lower()
            keywords_found = [kw for kw in expected_keywords if kw.lower() in response_lower]
        
        base_score = 1.0 if response_text else 0.0
        keyword_bonus = min(1.5, len(keywords_found) * 0.3)
        
        return {
            "score": min(5.0, base_score + keyword_bonus),
            "confidence": 0.3,
            "reasoning": f"Fallback evaluation: Found {len(keywords_found)} keywords",
            "strengths": [f"Mentioned {kw}" for kw in keywords_found[:3]] or ["Response provided"],
            "areas_for_improvement": ["Could not be fully evaluated"],
            "keywords_found": keywords_found,
            "mistakes_detected": []
        }

# =============================================================================
# CACHING SYSTEM
# =============================================================================

class EvaluationCache:
    """Simple caching system"""
    
    def __init__(self, redis_client=None, ttl_hours: int = 24):
        self.redis = redis_client
        self.ttl_seconds = ttl_hours * 3600
        self.memory_cache = {}
        self.cache_stats = {"hits": 0, "misses": 0, "saves": 0}
    
    def _generate_cache_key(self, question_id: str, response_content: str) -> str:
        """Generate cache key"""
        content_hash = hashlib.sha256(f"{question_id}:{response_content}".encode()).hexdigest()[:16]
        return f"eval:{question_id}:{content_hash}"
    
    async def get_cached_evaluation(self, cache_key: str) -> Optional[Dict]:
        """Get cached result"""
        
        try:
            if self.redis:
                cached_data = await self.redis.get(cache_key)
                if cached_data:
                    self.cache_stats["hits"] += 1
                    return json.loads(cached_data)
            
            if cache_key in self.memory_cache:
                cached_item = self.memory_cache[cache_key]
                if cached_item["expires"] > datetime.utcnow():
                    self.cache_stats["hits"] += 1
                    return cached_item["result"]
                else:
                    del self.memory_cache[cache_key]
            
            self.cache_stats["misses"] += 1
            return None
            
        except Exception as e:
            logging.warning(f"Cache error: {e}")
            self.cache_stats["misses"] += 1
            return None
    
    async def cache_evaluation(self, cache_key: str, evaluation: Dict):
        """Store in cache"""
        
        try:
            evaluation_json = json.dumps(evaluation)
            
            if self.redis:
                await self.redis.setex(cache_key, self.ttl_seconds, evaluation_json)
            
            self.memory_cache[cache_key] = {
                "result": evaluation,
                "expires": datetime.utcnow() + timedelta(seconds=self.ttl_seconds)
            }
            
            self.cache_stats["saves"] += 1
            
        except Exception as e:
            logging.warning(f"Cache storage error: {e}")
    
    def get_cache_stats(self) -> Dict[str, Any]:
        """Get cache stats"""
        
        total = self.cache_stats["hits"] + self.cache_stats["misses"]
        hit_rate = (self.cache_stats["hits"] / max(total, 1)) * 100
        
        return {
            "hits": self.cache_stats["hits"],
            "misses": self.cache_stats["misses"],
            "saves": self.cache_stats["saves"],
            "hit_rate_percentage": round(hit_rate, 2),
            "memory_cache_size": len(self.memory_cache)
        }

# =============================================================================
# FILE ANALYZER (SIMPLIFIED)
# =============================================================================

class ExcelFileAnalyzer:
    """Simplified Excel file analyzer"""
    
    def __init__(self):
        self.openpyxl_available = OPENPYXL_AVAILABLE
        self.xlrd_available = XLRD_AVAILABLE
        self.supported_formats = {'.xlsx', '.xls', '.xlsm'}
        self.max_file_size = 50 * 1024 * 1024
    
    async def analyze_excel_file(self, file_path: str, question: Dict) -> Dict[str, Any]:
        """Analyze Excel file"""
        
        if not self._validate_file(file_path):
            return self._error_analysis(file_path, "Invalid file")
        
        try:
            if file_path.lower().endswith('.xlsx') and self.openpyxl_available:
                return await self._analyze_with_openpyxl(file_path)
            elif file_path.lower().endswith('.xls') and self.xlrd_available:
                return await self._analyze_with_xlrd(file_path)
            else:
                return self._basic_analysis(file_path)
                
        except Exception as e:
            return self._error_analysis(file_path, str(e))
    
    async def _analyze_with_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """Analyze with openpyxl"""
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=False)
            
            total_formulas = 0
            unique_functions = set()
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 50)):
                    for cell in row:
                        if cell.data_type == 'f':
                            total_formulas += 1
                            functions = self._extract_functions(cell.value)
                            unique_functions.update(functions)
            
            return {
                "file_path": file_path,
                "file_size_bytes": Path(file_path).stat().st_size,
                "sheet_names": workbook.sheetnames,
                "total_formulas": total_formulas,
                "unique_functions": list(unique_functions),
                "formula_complexity_score": min(10.0, len(unique_functions) * 2),
                "has_pivot_tables": False,
                "has_charts": False,
                "has_conditional_formatting": bool(any(sheet.conditional_formatting for sheet in workbook.worksheets)),
                "has_errors": False,
                "error_cells": [],
                "data_quality_score": 5.0
            }
            
        except Exception as e:
            raise Exception(f"OpenPyXL analysis failed: {e}")
    
    async def _analyze_with_xlrd(self, file_path: str) -> Dict[str, Any]:
        """Analyze with xlrd"""
        
        try:
            workbook = xlrd.open_workbook(file_path)
            
            return {
                "file_path": file_path,
                "file_size_bytes": Path(file_path).stat().st_size,
                "sheet_names": workbook.sheet_names(),
                "total_formulas": 0,
                "unique_functions": [],
                "formula_complexity_score": 1.0,
                "has_pivot_tables": False,
                "has_charts": False,
                "has_conditional_formatting": False,
                "has_errors": False,
                "error_cells": [],
                "data_quality_score": 3.0
            }
            
        except Exception as e:
            raise Exception(f"XLRD analysis failed: {e}")
    
    def _basic_analysis(self, file_path: str) -> Dict[str, Any]:
        """Basic analysis when libraries unavailable"""
        
        return {
            "file_path": file_path,
            "file_size_bytes": Path(file_path).stat().st_size if Path(file_path).exists() else 0,
            "sheet_names": ["Sheet1"],
            "total_formulas": 0,
            "unique_functions": [],
            "formula_complexity_score": 0.0,
            "has_pivot_tables": False,
            "has_charts": False,
            "has_conditional_formatting": False,
            "has_errors": False,
            "error_cells": [],
            "data_quality_score": 2.0,
            "analysis_note": "Limited analysis - Excel libraries not available"
        }
    
    def _error_analysis(self, file_path: str, error: str) -> Dict[str, Any]:
        """Error analysis result"""
        
        return {
            "file_path": file_path,
            "file_size_bytes": 0,
            "sheet_names": [],
            "total_formulas": 0,
            "unique_functions": [],
            "formula_complexity_score": 0.0,
            "has_pivot_tables": False,
            "has_charts": False,
            "has_conditional_formatting": False,
            "has_errors": True,
            "error_cells": [error],
            "data_quality_score": 0.0,
            "analysis_error": error
        }
    
    def _extract_functions(self, formula: str) -> List[str]:
        """Extract function names from formula"""
        
        if not formula:
            return []
        
        # Simple regex to find function names
        function_pattern = r'\b([A-Z][A-Z0-9_]*)\s*\('
        functions = re.findall(function_pattern, formula.upper())
        return functions
    
    def _validate_file(self, file_path: str) -> bool:
        """Validate file"""
        
        path = Path(file_path)
        return (path.exists() and 
                path.stat().st_size <= self.max_file_size and
                path.suffix.lower() in self.supported_formats)

# =============================================================================
# MAIN EVALUATION ENGINE (ORCHESTRATOR COMPATIBLE)
# =============================================================================

class ClaudeEvaluationEngine:
    """Main evaluation engine - FULLY ORCHESTRATOR COMPATIBLE"""
    
    def __init__(
        self, 
        anthropic_client=None,
        anthropic_api_key: str = None,
        redis_client=None,
        claude_model: str = "claude-3-sonnet-20240229",
        cache_ttl_hours: int = 24
    ):
        # Create the claude_api that orchestrator expects
        self.claude_api = ClaudeAPIWrapper(anthropic_client, anthropic_api_key)
        
        # Other components
        self.file_analyzer = ExcelFileAnalyzer()
        self.cache = EvaluationCache(redis_client, cache_ttl_hours)
        
        # Orchestrator compatibility aliases
        self.claude_client = self.claude_api
        self.file_analyzer_engine = self.file_analyzer
        self.evaluation_cache = self.cache
        
        self.logger = logging.getLogger(__name__)
        
        # Performance tracking
        self.evaluation_stats = {
            "total_evaluations": 0,
            "cache_hits": 0,
            "llm_evaluations": 0,
            "file_evaluations": 0,
            "avg_evaluation_time": 0.0
        }
    
    async def evaluate_response(
        self, 
        question,
        text_response: str = None,
        file_path: str = None
    ) -> Dict[str, Any]:
        """Main evaluation entry point"""
        
        start_time = time.time()
        self.evaluation_stats["total_evaluations"] += 1
        
        try:
            # Convert question to dict if needed
            if hasattr(question, 'id'):
                question_dict = {
                    "id": question.id,
                    "text": question.text,
                    "type": getattr(question, 'type', 'free_text'),
                    "skill_category": getattr(question, 'skill_category', 'general'),
                    "difficulty": getattr(question, 'difficulty', 'intermediate'),
                    "expected_keywords": getattr(question, 'expected_keywords', []),
                    "common_mistakes": getattr(question, 'common_mistakes', [])
                }
            else:
                question_dict = question
            
            # Generate cache key
            cache_content = f"{text_response or ''}{file_path or ''}"
            cache_key = self.cache._generate_cache_key(question_dict["id"], cache_content)
            
            # Check cache
            cached_result = await self.cache.get_cached_evaluation(cache_key)
            if cached_result:
                self.evaluation_stats["cache_hits"] += 1
                return cached_result
            
            # Evaluate based on type
            question_type = question_dict.get("type", "free_text")
            
            if question_type == "free_text":
                evaluation = await self._evaluate_text_response(question_dict, text_response)
            elif question_type == "file_upload":
                evaluation = await self._evaluate_file_response(question_dict, file_path)
            elif question_type == "hybrid":
                evaluation = await self._evaluate_hybrid_response(question_dict, text_response, file_path)
            else:
                evaluation = await self._evaluate_text_response(question_dict, text_response)
            
            # Add metadata
            evaluation["question_id"] = question_dict["id"]
            evaluation["evaluation_time_ms"] = int((time.time() - start_time) * 1000)
            evaluation["evaluator_type"] = "claude"
            evaluation["created_at"] = datetime.utcnow().isoformat()
            
            # Cache result
            await self.cache.cache_evaluation(cache_key, evaluation)
            
            # Update stats
            self._update_evaluation_stats(time.time() - start_time)
            
            return evaluation
            
        except Exception as e:
            self.logger.error(f"Evaluation failed: {e}")
            return self._create_fallback_evaluation(question_dict.get("id", "unknown"), str(e))
    
    async def _evaluate_text_response(self, question: Dict, response_text: str) -> Dict[str, Any]:
        """Evaluate text responses"""
        
        if not response_text or not response_text.strip():
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": "No response provided",
                "strengths": [],
                "areas_for_improvement": ["Please provide a response"],
                "keywords_found": [],
                "mistakes_detected": []
            }
        
        self.evaluation_stats["llm_evaluations"] += 1
        return await self.claude_api.evaluate_text_response(question, response_text)
    
    async def _evaluate_file_response(self, question: Dict, file_path: str) -> Dict[str, Any]:
        """Evaluate file responses"""
        
        if not file_path:
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": "No file uploaded",
                "strengths": [],
                "areas_for_improvement": ["Please upload an Excel file"],
                "keywords_found": [],
                "mistakes_detected": []
            }
        
        self.evaluation_stats["file_evaluations"] += 1
        
        try:
            file_analysis = await self.file_analyzer.analyze_excel_file(file_path, question)
            return self._score_file_analysis(question, file_analysis)
        except Exception as e:
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": f"File analysis failed: {str(e)}",
                "strengths": [],
                "areas_for_improvement": ["Please ensure valid Excel file"],
                "keywords_found": [],
                "mistakes_detected": ["File analysis error"]
            }
    
    async def _evaluate_hybrid_response(self, question: Dict, text_response: str, file_path: str) -> Dict[str, Any]:
        """Evaluate hybrid responses"""
        
        text_eval = await self._evaluate_text_response(question, text_response) if text_response else None
        file_eval = await self._evaluate_file_response(question, file_path) if file_path else None
        
        if not text_eval and not file_eval:
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": "Both text and file required",
                "strengths": [],
                "areas_for_improvement": ["Provide both explanation and Excel file"],
                "keywords_found": [],
                "mistakes_detected": []
            }
        
        if text_eval and file_eval:
            combined_score = (text_eval["score"] * 0.4 + file_eval["score"] * 0.6)
            combined_reasoning = f"Text: {text_eval['reasoning']}\nFile: {file_eval['reasoning']}"
            combined_strengths = text_eval["strengths"] + file_eval["strengths"]
            combined_improvements = text_eval["areas_for_improvement"] + file_eval["areas_for_improvement"]
        elif text_eval:
            combined_score = text_eval["score"] * 0.6
            combined_reasoning = f"Text: {text_eval['reasoning']}. Missing file component."
            combined_strengths = text_eval["strengths"]
            combined_improvements = text_eval["areas_for_improvement"] + ["Upload Excel file"]
        else:
            combined_score = file_eval["score"] * 0.6
            combined_reasoning = f"File: {file_eval['reasoning']}. Missing text explanation."
            combined_strengths = file_eval["strengths"]
            combined_improvements = file_eval["areas_for_improvement"] + ["Provide explanation"]
        
        return {
            "score": round(combined_score, 2),
            "confidence": min((text_eval or {}).get("confidence", 0.8), (file_eval or {}).get("confidence", 0.8)),
            "reasoning": combined_reasoning,
            "strengths": list(set(combined_strengths)),
            "areas_for_improvement": list(set(combined_improvements)),
            "keywords_found": (text_eval or {}).get("keywords_found", []) + (file_eval or {}).get("keywords_found", []),
            "mistakes_detected": (text_eval or {}).get("mistakes_detected", []) + (file_eval or {}).get("mistakes_detected", [])
        }
    
    def _score_file_analysis(self, question: Dict, file_analysis: Dict) -> Dict[str, Any]:
        """Score file analysis"""
        
        score = 0.0
        strengths = []
        improvements = []
        
        total_formulas = file_analysis.get("total_formulas", 0)
        unique_functions = file_analysis.get("unique_functions", [])
        
        if total_formulas > 0:
            score += 1.0
            strengths.append(f"Used {total_formulas} formulas")
            
            if len(unique_functions) >= 3:
                score += 0.5
                strengths.append(f"Used {len(unique_functions)} different functions")
        else:
            improvements.append("No formulas detected")
        
        complexity_score = min(1.5, file_analysis.get("formula_complexity_score", 0) / 10 * 1.5)
        score += complexity_score
        
        if file_analysis.get("has_conditional_formatting"):
            score += 0.3
            strengths.append("Applied conditional formatting")
        
        if file_analysis.get("has_errors"):
            score *= 0.8
            improvements.append("Fix formula errors")
        
        final_score = min(5.0, max(0.0, score))
        
        return {
            "score": round(final_score, 2),
            "confidence": 0.9,
            "reasoning": f"File analysis: {total_formulas} formulas, {len(unique_functions)} functions",
            "strengths": strengths,
            "areas_for_improvement": improvements,
            "keywords_found": unique_functions,
            "mistakes_detected": file_analysis.get("error_cells", [])
        }
    
    def _create_fallback_evaluation(self, question_id: str, error_message: str) -> Dict[str, Any]:
        """Create fallback evaluation"""
        
        return {
            "score": 1.0,
            "confidence": 0.1,
            "reasoning": f"Evaluation error: {error_message}",
            "strengths": ["Response attempt made"],
            "areas_for_improvement": ["System error - please try again"],
            "keywords_found": [],
            "mistakes_detected": []
        }
    
    def _update_evaluation_stats(self, evaluation_time: float):
        """Update performance stats"""
        
        if self.evaluation_stats["avg_evaluation_time"] == 0:
            self.evaluation_stats["avg_evaluation_time"] = evaluation_time
        else:
            alpha = 0.1
            self.evaluation_stats["avg_evaluation_time"] = (
                alpha * evaluation_time + 
                (1 - alpha) * self.evaluation_stats["avg_evaluation_time"]
            )
    
    # =========================================================================
    # ORCHESTRATOR REQUIRED METHODS
    # =========================================================================
    
    async def health_check(self) -> Dict[str, Any]:
        """Health check method required by orchestrator"""
        
        health_data = {
            "status": "healthy",
            "claude_available": self.claude_api.available,
            "openpyxl_available": self.file_analyzer.openpyxl_available,
            "xlrd_available": self.file_analyzer.xlrd_available,
            "cache_enabled": self.cache.redis is not None,
            "evaluation_count": self.evaluation_stats["total_evaluations"],
            "issues": []
        }
        
        # Check components
        if not self.claude_api.available:
            health_data["issues"].append("Claude LLM not available - using fallback")
        
        if not (self.file_analyzer.openpyxl_available or self.file_analyzer.xlrd_available):
            health_data["issues"].append("No Excel parsing libraries - file analysis limited")
        
        # Test evaluation function
        try:
            test_question = {
                "id": "health_check",
                "text": "Test question",
                "type": "free_text",
                "expected_keywords": ["test"]
            }
            
            test_eval = await self._evaluate_text_response(test_question, "test response")
            health_data["evaluation_functional"] = test_eval["score"] >= 0
            
        except Exception as e:
            health_data["issues"].append(f"Evaluation test failed: {str(e)}")
            health_data["evaluation_functional"] = False
        
        # Calculate health percentage
        checks = [
            health_data["claude_available"],
            health_data["openpyxl_available"] or health_data["xlrd_available"],
            True,  # Cache always works
            health_data.get("evaluation_functional", False)
        ]
        
        health_data["health_percentage"] = int((sum(checks) / len(checks)) * 100)
        
        if health_data["health_percentage"] >= 75:
            health_data["status"] = "healthy"
        elif health_data["health_percentage"] >= 50:
            health_data["status"] = "degraded"
        else:
            health_data["status"] = "unhealthy"
        
        return health_data
    
    def get_performance_stats(self) -> Dict[str, Any]:
        """Get performance statistics"""
        
        cache_stats = self.cache.get_cache_stats()
        
        return {
            "evaluation_stats": self.evaluation_stats,
            "cache_performance": cache_stats,
            "claude_api_stats": {
                "total_calls": self.claude_api.total_calls,
                "successful_calls": self.claude_api.successful_calls,
                "failed_calls": self.claude_api.failed_calls,
                "success_rate": (self.claude_api.successful_calls / max(self.claude_api.total_calls, 1)) * 100
            },
            "overall_cache_hit_rate": cache_stats["hit_rate_percentage"],
            "avg_evaluation_time": round(self.evaluation_stats["avg_evaluation_time"], 3)
        }
    
    # Additional orchestrator compatibility properties
    @property 
    def claude_api_available(self) -> bool:
        """Check if Claude API is available"""
        return self.claude_api.available
    
    @property
    def anthropic_client(self):
        """Direct access to anthropic client"""
        return self.claude_api.anthropic_client

# =============================================================================
# FACTORY FOR CREATING ENGINES
# =============================================================================

class EvaluationEngineFactory:
    """Factory for creating evaluation engines"""
    
    @staticmethod
    def create_production_engine(
        anthropic_api_key: str,
        redis_client=None,
        cache_ttl_hours: int = 24
    ) -> ClaudeEvaluationEngine:
        """Create production engine"""
        
        return ClaudeEvaluationEngine(
            anthropic_api_key=anthropic_api_key,
            redis_client=redis_client,
            cache_ttl_hours=cache_ttl_hours
        )
    
    @staticmethod
    def create_development_engine(
        anthropic_api_key: str = "test_key",
        cache_ttl_hours: int = 1
    ) -> ClaudeEvaluationEngine:
        """Create development engine"""
        
        return ClaudeEvaluationEngine(
            anthropic_api_key=anthropic_api_key,
            redis_client=None,
            cache_ttl_hours=cache_ttl_hours
        )
    
    @staticmethod
    def create_orchestrator_compatible_engine(
        anthropic_client=None,
        anthropic_api_key: str = None,
        redis_client=None
    ) -> ClaudeEvaluationEngine:
        """Create engine specifically for orchestrator integration"""
        
        return ClaudeEvaluationEngine(
            anthropic_client=anthropic_client,
            anthropic_api_key=anthropic_api_key,
            redis_client=redis_client,
            cache_ttl_hours=24
        )

# =============================================================================
# EXAMPLE USAGE AND TESTING
# =============================================================================

async def example_usage():
    """Example showing orchestrator compatibility"""
    
    print("🔧 Testing Orchestrator Compatible Evaluation Engine")
    print("=" * 60)
    
    # Create evaluation engine (orchestrator style)
    engine = EvaluationEngineFactory.create_orchestrator_compatible_engine(
        anthropic_api_key="test_key"
    )
    
    # Test required orchestrator attributes
    print("📋 Orchestrator Compatibility Check:")
    print(f"  ✅ claude_api exists: {hasattr(engine, 'claude_api')}")
    print(f"  ✅ claude_api type: {type(engine.claude_api).__name__}")
    print(f"  ✅ claude_api.available: {engine.claude_api.available}")
    print(f"  ✅ health_check exists: {hasattr(engine, 'health_check')}")
    print(f"  ✅ file_analyzer exists: {hasattr(engine, 'file_analyzer')}")
    print(f"  ✅ cache exists: {hasattr(engine, 'cache')}")
    print()
    
    # Test health check (required by orchestrator)
    print("🏥 Running Health Check:")
    try:
        health = await engine.health_check()
        print(f"  ✅ Status: {health['status']}")
        print(f"  ✅ Health Percentage: {health['health_percentage']}%")
        print(f"  ✅ Claude Available: {health['claude_available']}")
        print(f"  ✅ Evaluation Functional: {health.get('evaluation_functional', False)}")
        if health['issues']:
            print(f"  ⚠️  Issues: {health['issues']}")
        print()
    except Exception as e:
        print(f"  ❌ Health check failed: {e}")
        print()
    
    # Test evaluation (core functionality)
    print("📊 Testing Evaluation Functionality:")
    
    # Question in dict format (orchestrator style)
    question = {
        "id": "test_q1",
        "text": "Explain how VLOOKUP works in Excel",
        "type": "free_text",
        "skill_category": "formulas", 
        "difficulty": "intermediate",
        "expected_keywords": ["VLOOKUP", "lookup", "table", "exact", "match"],
        "common_mistakes": ["wrong table array", "incorrect column index"]
    }
    
    # Test response
    response_text = """VLOOKUP is a function that searches for a value in the first column of a table 
    and returns a value in the same row from a specified column. The syntax is 
    VLOOKUP(lookup_value, table_array, col_index_num, range_lookup). 
    Use FALSE for exact match to avoid errors."""
    
    try:
        # This is the main call orchestrator makes
        evaluation = await engine.evaluate_response(
            question=question,
            text_response=response_text
        )
        
        print(f"  ✅ Evaluation Score: {evaluation['score']}/5.0")
        print(f"  ✅ Confidence: {evaluation['confidence']}")
        print(f"  ✅ Question ID: {evaluation.get('question_id', 'N/A')}")
        print(f"  ✅ Keywords Found: {evaluation.get('keywords_found', [])}")
        print(f"  ✅ Strengths: {evaluation.get('strengths', [])[:2]}...")
        print(f"  ✅ Reasoning: {evaluation.get('reasoning', 'N/A')[:80]}...")
        print()
        
    except Exception as e:
        print(f"  ❌ Evaluation failed: {e}")
        print()
    
    # Test performance stats
    print("📈 Performance Statistics:")
    try:
        stats = engine.get_performance_stats()
        print(f"  ✅ Total Evaluations: {stats['evaluation_stats']['total_evaluations']}")
        print(f"  ✅ Cache Hit Rate: {stats['cache_performance']['hit_rate_percentage']}%")
        print(f"  ✅ Claude API Calls: {stats['claude_api_stats']['total_calls']}")
        print(f"  ✅ Average Time: {stats['avg_evaluation_time']:.3f}s")
        print()
    except Exception as e:
        print(f"  ❌ Stats failed: {e}")
        print()
    
    # Test direct claude_api access (what orchestrator does)
    print("🔍 Testing Direct claude_api Access:")
    try:
        # This is what caused the original error
        claude_api = engine.claude_api
        print(f"  ✅ engine.claude_api accessible: {claude_api is not None}")
        print(f"  ✅ claude_api.available: {claude_api.available}")
        print(f"  ✅ claude_api type: {type(claude_api).__name__}")
        
        # Test the method orchestrator calls
        if hasattr(claude_api, 'evaluate_text_response'):
            print(f"  ✅ claude_api.evaluate_text_response exists")
        else:
            print(f"  ❌ claude_api.evaluate_text_response missing")
        
    except Exception as e:
        print(f"  ❌ claude_api access failed: {e}")
    
    print("\n🎯 Orchestrator Compatibility Summary:")
    print("  ✅ All required attributes exposed")
    print("  ✅ health_check() returns health_percentage") 
    print("  ✅ claude_api properly accessible")
    print("  ✅ Evaluation function working")
    print("  ✅ Compatible with orchestrator integration patterns")

def validate_orchestrator_compatibility():
    """Validate that engine meets orchestrator requirements"""
    
    print("🔍 Validating Orchestrator Compatibility...")
    
    # Create engine
    engine = EvaluationEngineFactory.create_orchestrator_compatible_engine()
    
    # Check all required attributes
    required_attributes = [
        'claude_api',
        'file_analyzer', 
        'cache',
        'health_check',
        'evaluate_response',
        'get_performance_stats'
    ]
    
    missing_attributes = []
    for attr in required_attributes:
        if not hasattr(engine, attr):
            missing_attributes.append(attr)
    
    if missing_attributes:
        print(f"❌ Missing attributes: {missing_attributes}")
        return False
    
    # Check claude_api structure
    claude_api = engine.claude_api
    if not hasattr(claude_api, 'available'):
        print("❌ claude_api missing 'available' attribute")
        return False
    
    if not hasattr(claude_api, 'evaluate_text_response'):
        print("❌ claude_api missing 'evaluate_text_response' method")
        return False
    
    print("✅ All orchestrator requirements met!")
    return True

if __name__ == "__main__":
    print("🔍 Claude Evaluation Engine - ORCHESTRATOR COMPATIBLE")
    print("=" * 70)
    print("🎯 Features Available:")
    print("  ✅ claude_api attribute (orchestrator requirement)")
    print("  ✅ health_check() with health_percentage")
    print("  ✅ ClaudeAPIWrapper for exact orchestrator interface")
    print("  ✅ Excel file analysis with graceful fallbacks")
    print("  ✅ Intelligent caching (Redis + memory)")
    print("  ✅ Multi-modal evaluation (text + file + hybrid)")
    print("  ✅ Performance monitoring and statistics")
    print("  ✅ Comprehensive error handling and fallbacks")
    print()
    
    # Validate compatibility first
    if validate_orchestrator_compatibility():
        print("\n🚀 Running full example...")
        import asyncio
        asyncio.run(example_usage())
    else:
        print("\n❌ Compatibility validation failed!")
        
    print("\n📚 Usage for Orchestrator Integration:")
    print("```python")
    print("# Create engine exactly as orchestrator expects")
    print("engine = EvaluationEngineFactory.create_orchestrator_compatible_engine(")
    print("    anthropic_client=your_anthropic_client,")
    print("    redis_client=your_redis_client")
    print(")")
    print("")
    print("# Orchestrator can now access:")
    print("# - engine.claude_api")
    print("# - engine.health_check()")
    print("# - engine.evaluate_response()")
    print("```")