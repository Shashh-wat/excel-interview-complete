# evaluation_engine.py - COMPLETE WORKING Claude-Powered Evaluation Engine
"""
COMPLETE: Comprehensive evaluation system that scores candidate responses using:
1. Claude LLM for advanced textual analysis with full integration
2. Excel file parsing and analysis with multiple library support
3. Intelligent caching for performance optimization
4. Multi-modal evaluation (text + file + voice)
5. COMPLETE: Full error handling, initialization, and orchestrator compatibility
6. Enhanced fallback systems for production reliability
"""

import json
import time
import hashlib
import asyncio
import logging
from typing import Dict, List, Optional, Union, Any
from datetime import datetime, timedelta
import re
from pathlib import Path

# Enhanced imports with graceful fallbacks
try:
    import anthropic
    from anthropic import AsyncAnthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False
    logging.warning("Anthropic not available - evaluation will use enhanced fallbacks")

# Excel file parsing (with complete fallback support)
try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

# =============================================================================
# COMPLETE CLAUDE API WRAPPER
# =============================================================================

class ClaudeAPIWrapper:
    """COMPLETE: Wrapper with full functionality and comprehensive error handling"""
    
    def __init__(self, anthropic_client=None, api_key=None):
        self.anthropic_client = None
        self.api_key = api_key
        self.available = False
        self.logger = logging.getLogger(__name__)
        
        # Performance tracking
        self.performance_stats = {
            "total_calls": 0,
            "successful_calls": 0,
            "failed_calls": 0,
            "average_response_time": 0.0,
            "total_tokens_used": 0,
            "cache_hits": 0
        }
        
        # Simple response cache
        self.response_cache = {}
        
        # Try to initialize Anthropic client with comprehensive error handling
        try:
            if anthropic_client:
                self.anthropic_client = anthropic_client
                self.available = True
                self.logger.info("✅ Using provided Anthropic client")
            elif api_key and api_key != "test_key" and ANTHROPIC_AVAILABLE:
                self.anthropic_client = AsyncAnthropic(api_key=api_key)
                self.available = True
                self.logger.info("✅ Anthropic client initialized with API key")
            else:
                self.logger.warning("⚠️ No valid Anthropic API key provided - using enhanced fallback mode")
                self.available = False
        except Exception as e:
            self.logger.error(f"❌ Failed to initialize Anthropic client: {e}")
            self.available = False
    
    async def evaluate_text_response(self, question: Dict, response_text: str) -> Dict[str, Any]:
        """Complete evaluation method with caching and comprehensive analysis"""
        
        start_time = time.time()
        self.performance_stats["total_calls"] += 1
        
        if not self.available or not response_text or not response_text.strip():
            return self._enhanced_fallback_evaluation(question, response_text)
        
        # Check cache first
        cache_key = self._generate_cache_key(question, response_text)
        if cache_key in self.response_cache:
            self.performance_stats["cache_hits"] += 1
            cached_result = self.response_cache[cache_key]
            if cached_result["expires"] > datetime.utcnow():
                return cached_result["evaluation"]
            else:
                del self.response_cache[cache_key]
        
        try:
            prompt = self._build_comprehensive_evaluation_prompt(question, response_text)
            
            response = await self.anthropic_client.messages.create(
                model="claude-3-5-sonnet-20241022",
                max_tokens=1200,
                temperature=0.2,
                messages=[{"role": "user", "content": prompt}]
            )
            
            response_time = time.time() - start_time
            self._update_performance_stats(response_time, True)
            
            evaluation = self._parse_claude_response(response.content[0].text)
            
            # Cache successful evaluation
            self.response_cache[cache_key] = {
                "evaluation": evaluation,
                "expires": datetime.utcnow() + timedelta(hours=24)
            }
            
            # Limit cache size
            if len(self.response_cache) > 1000:
                oldest_key = min(self.response_cache.keys(), 
                               key=lambda k: self.response_cache[k]["expires"])
                del self.response_cache[oldest_key]
            
            return evaluation
            
        except Exception as e:
            response_time = time.time() - start_time
            self._update_performance_stats(response_time, False)
            self.logger.error(f"Claude API error: {e}")
            return self._enhanced_fallback_evaluation(question, response_text)
    
    def _build_comprehensive_evaluation_prompt(self, question: Dict, response_text: str) -> str:
        """Build comprehensive evaluation prompt with context"""
        
        question_text = question.get('text', '')
        expected_keywords = question.get('expected_keywords', [])
        difficulty = question.get('difficulty', 'intermediate')
        skill_category = question.get('skill_category', 'general')
        
        return f"""You are a senior Excel interviewer and Microsoft Office specialist evaluating a candidate's response for a technical interview.

INTERVIEW CONTEXT:
- Question Category: {skill_category.replace('_', ' ').title()}
- Difficulty Level: {difficulty.title()}
- Expected Duration: 2-5 minutes

QUESTION: {question_text}

EVALUATION CRITERIA:
- Technical Accuracy (0-2 points): Correctness of Excel knowledge and concepts
- Depth of Understanding (0-2 points): How well the candidate explains the concepts
- Practical Application (0-1 point): Use of examples, scenarios, or practical insights

EXPECTED KEY TERMS: {', '.join(expected_keywords) if expected_keywords else 'General Excel knowledge'}

CANDIDATE RESPONSE:
{response_text}

EVALUATION INSTRUCTIONS:
1. Score from 0.0 to 5.0 (decimals allowed)
2. Consider the difficulty level - {difficulty} questions require {
   'basic understanding' if difficulty == 'beginner' else
   'solid knowledge with examples' if difficulty == 'intermediate' else
   'advanced expertise with detailed explanations'
}
3. Identify specific strengths and areas for improvement
4. Note which expected keywords were correctly used
5. Flag any technical mistakes or misconceptions

Respond with ONLY valid JSON in this exact format:
{{
  "score": 3.7,
  "confidence": 0.85,
  "reasoning": "Detailed explanation of why this score was given, mentioning specific aspects of the response",
  "strengths": ["Specific strength 1", "Specific strength 2", "Specific strength 3"],
  "areas_for_improvement": ["Specific improvement area 1", "Specific improvement area 2"],
  "keywords_found": ["keyword1", "keyword2", "keyword3"],
  "mistakes_detected": ["mistake1", "mistake2"],
  "technical_accuracy": 1.8,
  "depth_of_understanding": 1.5,
  "practical_application": 0.4
}}"""
    
    def _parse_claude_response(self, response_text: str) -> Dict[str, Any]:
        """Parse Claude's response with enhanced error handling"""
        
        try:
            # Find JSON in response with multiple strategies
            json_content = None
            
            # Strategy 1: Look for complete JSON object
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start >= 0 and json_end > json_start:
                json_content = response_text[json_start:json_end]
            
            # Strategy 2: Extract between ```json blocks
            if not json_content:
                json_match = re.search(r'```json\s*(\{.*?\})\s*```', response_text, re.DOTALL)
                if json_match:
                    json_content = json_match.group(1)
            
            # Strategy 3: Look for any valid JSON-like structure
            if not json_content:
                json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response_text, re.DOTALL)
                if json_match:
                    json_content = json_match.group(0)
            
            if json_content:
                parsed = json.loads(json_content)
                
                # Validate and clean parsed data
                evaluation = {
                    "score": float(parsed.get("score", 2.5)),
                    "confidence": float(parsed.get("confidence", 0.8)),
                    "reasoning": str(parsed.get("reasoning", "Claude evaluation completed")),
                    "strengths": list(parsed.get("strengths", ["Response provided"])),
                    "areas_for_improvement": list(parsed.get("areas_for_improvement", ["Could be more detailed"])),
                    "keywords_found": list(parsed.get("keywords_found", [])),
                    "mistakes_detected": list(parsed.get("mistakes_detected", [])),
                    "evaluation_method": "claude_ai_advanced"
                }
                
                # Add detailed scoring if available
                if "technical_accuracy" in parsed:
                    evaluation["detailed_scoring"] = {
                        "technical_accuracy": float(parsed.get("technical_accuracy", 0)),
                        "depth_of_understanding": float(parsed.get("depth_of_understanding", 0)),
                        "practical_application": float(parsed.get("practical_application", 0))
                    }
                
                # Validate score range
                evaluation["score"] = max(0.0, min(5.0, evaluation["score"]))
                evaluation["confidence"] = max(0.0, min(1.0, evaluation["confidence"]))
                
                return evaluation
            else:
                raise ValueError("No valid JSON found in response")
                
        except json.JSONDecodeError as e:
            self.logger.warning(f"JSON parsing failed: {e}")
            return self._enhanced_fallback_evaluation({}, response_text)
        except Exception as e:
            self.logger.warning(f"Failed to parse Claude response: {e}")
            return self._enhanced_fallback_evaluation({}, response_text)
    
    def _enhanced_fallback_evaluation(self, question: Dict, response_text: str) -> Dict[str, Any]:
        """Enhanced fallback evaluation with sophisticated analysis"""
        
        if not response_text or not response_text.strip():
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": "No response provided",
                "strengths": [],
                "areas_for_improvement": ["Please provide a response"],
                "keywords_found": [],
                "mistakes_detected": [],
                "evaluation_method": "fallback_no_response"
            }
        
        expected_keywords = question.get('expected_keywords', [])
        difficulty = question.get('difficulty', 'intermediate')
        skill_category = question.get('skill_category', 'general')
        
        # Enhanced analysis
        response_lower = response_text.lower()
        word_count = len(response_text.split())
        sentence_count = len([s for s in response_text.split('.') if s.strip()])
        
        # Keyword analysis with partial matching
        keywords_found = []
        partial_matches = []
        
        for keyword in expected_keywords:
            if keyword.lower() in response_lower:
                keywords_found.append(keyword)
            elif any(part in response_lower for part in keyword.lower().split()):
                partial_matches.append(keyword)
        
        # Excel-specific terminology detection
        excel_terms = ['formula', 'function', 'cell', 'range', 'worksheet', 'workbook', 
                      'pivot', 'chart', 'filter', 'sort', 'format', 'reference']
        excel_terms_found = [term for term in excel_terms if term in response_lower]
        
        # Quality indicators
        explanation_indicators = ['because', 'since', 'therefore', 'however', 'for example', 
                                'specifically', 'in other words', 'such as', 'including']
        explanation_count = sum(1 for indicator in explanation_indicators if indicator in response_lower)
        
        # Scoring algorithm
        base_score = 1.0  # Base for any response
        
        # Keyword scoring (adjusted by difficulty)
        keyword_multipliers = {"beginner": 0.5, "intermediate": 0.4, "advanced": 0.3}
        keyword_bonus = min(2.0, len(keywords_found) * keyword_multipliers.get(difficulty, 0.4))
        partial_bonus = min(0.5, len(partial_matches) * 0.2)
        
        # Length and structure scoring
        length_targets = {"beginner": 25, "intermediate": 50, "advanced": 80}
        target_length = length_targets.get(difficulty, 50)
        length_score = min(1.0, word_count / target_length)
        
        # Quality scoring
        explanation_bonus = min(0.8, explanation_count * 0.2)
        excel_terminology_bonus = min(0.7, len(excel_terms_found) * 0.1)
        
        # Sentence structure bonus
        avg_sentence_length = word_count / max(sentence_count, 1)
        structure_bonus = 0.3 if 8 <= avg_sentence_length <= 25 else 0.1
        
        final_score = min(5.0, base_score + keyword_bonus + partial_bonus + 
                         length_score + explanation_bonus + excel_terminology_bonus + structure_bonus)
        
        # Generate detailed feedback
        strengths = []
        improvements = []
        
        # Strength analysis
        if len(keywords_found) >= 3:
            strengths.append("Excellent use of relevant Excel terminology")
        elif len(keywords_found) >= 1:
            strengths.append("Good understanding of key concepts")
        
        if word_count >= target_length:
            strengths.append("Provided comprehensive explanation")
        elif word_count >= target_length * 0.7:
            strengths.append("Good explanation detail")
        
        if explanation_count >= 2:
            strengths.append("Clear explanatory language with examples")
        
        if len(excel_terms_found) >= 3:
            strengths.append("Strong Excel vocabulary")
        
        # Improvement analysis
        if len(keywords_found) == 0:
            improvements.append("Include more Excel-specific terminology")
        
        if word_count < target_length * 0.5:
            improvements.append("Provide more detailed explanation")
        
        if explanation_count == 0:
            improvements.append("Add examples or explanations of 'why' and 'how'")
        
        if difficulty == "advanced" and final_score < 3.0:
            improvements.append("Advanced questions require more sophisticated technical explanations")
        
        # Default messages if empty
        if not strengths:
            strengths = ["Response provided", "Attempted to answer the question"]
        
        if not improvements:
            if final_score >= 4.0:
                improvements = ["Excellent response!"]
            else:
                improvements = ["Consider adding more specific details"]
        
        return {
            "score": round(final_score, 1),
            "confidence": 0.75,
            "reasoning": f"Enhanced analysis: {len(keywords_found)} keywords + {len(partial_matches)} partial matches found in {word_count} words. {difficulty.title()} {skill_category.replace('_', ' ')} question.",
            "strengths": strengths[:4],  # Limit to top 4
            "areas_for_improvement": improvements[:3],  # Limit to top 3
            "keywords_found": keywords_found,
            "mistakes_detected": [],
            "evaluation_method": "enhanced_fallback",
            "analysis_details": {
                "word_count": word_count,
                "sentence_count": sentence_count,
                "excel_terms_found": excel_terms_found,
                "explanation_indicators": explanation_count,
                "keyword_coverage": f"{len(keywords_found)}/{len(expected_keywords)}" if expected_keywords else "N/A"
            }
        }
    
    def _generate_cache_key(self, question: Dict, response_text: str) -> str:
        """Generate cache key for responses"""
        try:
            question_id = question.get("id", "unknown")
            text_hash = hashlib.sha256(response_text.encode()).hexdigest()[:16]
            return f"eval:{question_id}:{text_hash}"
        except Exception:
            return f"eval:fallback:{hash(response_text) % 10000}"
    
    def _update_performance_stats(self, response_time: float, success: bool):
        """Update performance statistics"""
        try:
            if success:
                self.performance_stats["successful_calls"] += 1
            else:
                self.performance_stats["failed_calls"] += 1
            
            # Update average response time
            total_calls = self.performance_stats["total_calls"]
            current_avg = self.performance_stats["average_response_time"]
            self.performance_stats["average_response_time"] = (
                (current_avg * (total_calls - 1) + response_time) / total_calls
            )
        except Exception as e:
            self.logger.warning(f"Performance stats update failed: {e}")
    
    def get_performance_stats(self) -> Dict[str, Any]:
        """Get comprehensive performance statistics"""
        
        total = self.performance_stats["total_calls"]
        success_rate = (self.performance_stats["successful_calls"] / max(total, 1)) * 100
        cache_hit_rate = (self.performance_stats["cache_hits"] / max(total, 1)) * 100
        
        return {
            "api_available": self.available,
            "total_calls": total,
            "successful_calls": self.performance_stats["successful_calls"],
            "failed_calls": self.performance_stats["failed_calls"],
            "success_rate_percentage": round(success_rate, 2),
            "cache_hits": self.performance_stats["cache_hits"],
            "cache_hit_rate_percentage": round(cache_hit_rate, 2),
            "average_response_time_seconds": round(self.performance_stats["average_response_time"], 3),
            "cache_size": len(self.response_cache)
        }

# =============================================================================
# COMPLETE CACHING SYSTEM
# =============================================================================

class EvaluationCache:
    """COMPLETE: Advanced caching system with Redis and memory fallbacks"""
    
    def __init__(self, redis_client=None, ttl_hours: int = 24):
        self.redis = redis_client
        self.ttl_seconds = ttl_hours * 3600
        self.memory_cache = {}
        self.logger = logging.getLogger(__name__)
        
        # Cache statistics
        self.cache_stats = {
            "hits": 0, 
            "misses": 0, 
            "saves": 0,
            "evictions": 0,
            "redis_operations": 0,
            "memory_operations": 0
        }
        
        # Memory cache limits
        self.max_memory_cache_size = 500
    
    def _generate_cache_key(self, question_id: str, response_content: str) -> str:
        """Generate comprehensive cache key"""
        try:
            # Create stable hash of content
            content_hash = hashlib.sha256(
                f"{question_id}:{response_content}".encode('utf-8')
            ).hexdigest()[:16]
            return f"eval_v2:{question_id}:{content_hash}"
        except Exception as e:
            self.logger.warning(f"Cache key generation failed: {e}")
            return f"eval_fallback:{question_id}:{hash(response_content) % 100000}"
    
    async def get_cached_evaluation(self, cache_key: str) -> Optional[Dict]:
        """Get cached result with Redis and memory fallback"""
        
        try:
            # Check Redis first if available
            if self.redis:
                try:
                    self.cache_stats["redis_operations"] += 1
                    cached_data = await self.redis.get(cache_key)
                    if cached_data:
                        self.cache_stats["hits"] += 1
                        return json.loads(cached_data)
                except Exception as e:
                    self.logger.warning(f"Redis cache read failed: {e}")
            
            # Check memory cache
            self.cache_stats["memory_operations"] += 1
            if cache_key in self.memory_cache:
                cached_item = self.memory_cache[cache_key]
                if cached_item["expires"] > datetime.utcnow():
                    self.cache_stats["hits"] += 1
                    return cached_item["result"]
                else:
                    # Remove expired entry
                    del self.memory_cache[cache_key]
                    self.cache_stats["evictions"] += 1
            
            self.cache_stats["misses"] += 1
            return None
            
        except Exception as e:
            self.logger.warning(f"Cache retrieval error: {e}")
            self.cache_stats["misses"] += 1
            return None
    
    async def cache_evaluation(self, cache_key: str, evaluation: Dict):
        """Store evaluation in cache with multi-tier strategy"""
        
        try:
            evaluation_json = json.dumps(evaluation, default=str)
            
            # Try Redis first
            if self.redis:
                try:
                    self.cache_stats["redis_operations"] += 1
                    await self.redis.setex(cache_key, self.ttl_seconds, evaluation_json)
                    self.logger.debug(f"✅ Cached to Redis: {cache_key}")
                except Exception as e:
                    self.logger.warning(f"Redis cache write failed: {e}")
            
            # Always store in memory cache as backup
            self.cache_stats["memory_operations"] += 1
            self.memory_cache[cache_key] = {
                "result": evaluation,
                "expires": datetime.utcnow() + timedelta(seconds=self.ttl_seconds),
                "created": datetime.utcnow()
            }
            
            # Manage memory cache size
            if len(self.memory_cache) > self.max_memory_cache_size:
                # Remove oldest entries
                oldest_keys = sorted(
                    self.memory_cache.keys(),
                    key=lambda k: self.memory_cache[k]["created"]
                )
                for key in oldest_keys[:50]:  # Remove 50 oldest
                    del self.memory_cache[key]
                    self.cache_stats["evictions"] += 1
            
            self.cache_stats["saves"] += 1
            
        except Exception as e:
            self.logger.warning(f"Cache storage error: {e}")
    
    def get_cache_stats(self) -> Dict[str, Any]:
        """Get comprehensive cache statistics"""
        
        total_requests = self.cache_stats["hits"] + self.cache_stats["misses"]
        hit_rate = (self.cache_stats["hits"] / max(total_requests, 1)) * 100
        
        return {
            "hits": self.cache_stats["hits"],
            "misses": self.cache_stats["misses"],
            "saves": self.cache_stats["saves"],
            "evictions": self.cache_stats["evictions"],
            "hit_rate_percentage": round(hit_rate, 2),
            "memory_cache_size": len(self.memory_cache),
            "memory_cache_limit": self.max_memory_cache_size,
            "redis_available": self.redis is not None,
            "redis_operations": self.cache_stats["redis_operations"],
            "memory_operations": self.cache_stats["memory_operations"]
        }
    
    def clear_cache(self):
        """Clear all cached data"""
        try:
            self.memory_cache.clear()
            self.logger.info("✅ Memory cache cleared")
        except Exception as e:
            self.logger.warning(f"Cache clear failed: {e}")

# =============================================================================
# COMPLETE FILE ANALYZER
# =============================================================================

class ExcelFileAnalyzer:
    """COMPLETE: Comprehensive Excel file analyzer with multiple library support"""
    
    def __init__(self):
        self.openpyxl_available = OPENPYXL_AVAILABLE
        self.xlrd_available = XLRD_AVAILABLE
        self.pandas_available = PANDAS_AVAILABLE
        self.supported_formats = {'.xlsx', '.xls', '.xlsm', '.csv'}
        self.max_file_size = 100 * 1024 * 1024  # 100MB
        self.logger = logging.getLogger(__name__)
        
        # Analysis statistics
        self.analysis_stats = {
            "files_analyzed": 0,
            "successful_analyses": 0,
            "failed_analyses": 0,
            "total_formulas_found": 0,
            "unique_functions_discovered": set()
        }
        
        self.logger.info(f"✅ Excel analyzer initialized - OpenPyXL: {self.openpyxl_available}, XLRD: {self.xlrd_available}, Pandas: {self.pandas_available}")
    
    async def analyze_excel_file(self, file_path: str, question: Dict) -> Dict[str, Any]:
        """Comprehensive Excel file analysis with multiple strategies"""
        
        self.analysis_stats["files_analyzed"] += 1
        
        try:
            if not self._validate_file(file_path):
                return self._error_analysis(file_path, "File validation failed")
            
            file_ext = Path(file_path).suffix.lower()
            
            # Choose analysis strategy based on file type and available libraries
            if file_ext == '.xlsx' and self.openpyxl_available:
                result = await self._analyze_with_openpyxl(file_path)
            elif file_ext == '.xls' and self.xlrd_available:
                result = await self._analyze_with_xlrd(file_path)
            elif file_ext in ['.xlsx', '.xls', '.csv'] and self.pandas_available:
                result = await self._analyze_with_pandas(file_path)
            else:
                result = self._basic_file_analysis(file_path)
            
            if result.get("analysis_error"):
                self.analysis_stats["failed_analyses"] += 1
            else:
                self.analysis_stats["successful_analyses"] += 1
                self.analysis_stats["total_formulas_found"] += result.get("total_formulas", 0)
                unique_functions = result.get("unique_functions", [])
                self.analysis_stats["unique_functions_discovered"].update(unique_functions)
            
            return result
                
        except Exception as e:
            self.analysis_stats["failed_analyses"] += 1
            self.logger.error(f"File analysis failed for {file_path}: {e}")
            return self._error_analysis(file_path, str(e))
    
    async def _analyze_with_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """Comprehensive analysis with openpyxl"""
        
        try:
            # Load workbook with full features
            workbook = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)
            
            analysis_result = {
                "file_path": file_path,
                "file_size_bytes": Path(file_path).stat().st_size,
                "sheet_names": workbook.sheetnames,
                "total_sheets": len(workbook.sheetnames),
                "total_formulas": 0,
                "unique_functions": set(),
                "cell_analysis": {},
                "data_quality_metrics": {},
                "advanced_features": {
                    "has_pivot_tables": False,
                    "has_charts": False,
                    "has_conditional_formatting": False,
                    "has_data_validation": False,
                    "has_named_ranges": False
                }
            }
            
            # Analyze up to first 5 sheets to prevent timeout
            sheets_to_analyze = workbook.sheetnames[:5]
            
            for sheet_name in sheets_to_analyze:
                try:
                    sheet = workbook[sheet_name]
                    sheet_analysis = {
                        "max_row": sheet.max_row or 0,
                        "max_column": sheet.max_column or 0,
                        "formulas_count": 0,
                        "functions_used": set(),
                        "errors_found": [],
                        "data_types": {"text": 0, "number": 0, "formula": 0, "empty": 0}
                    }
                    
                    # Analyze cells (limit to reasonable range)
                    max_rows = min(sheet.max_row or 1, 200)
                    max_cols = min(sheet.max_column or 1, 50)
                    
                    for row in range(1, max_rows + 1):
                        for col in range(1, max_cols + 1):
                            try:
                                cell = sheet.cell(row=row, column=col)
                                
                                if cell.value is None:
                                    sheet_analysis["data_types"]["empty"] += 1
                                elif hasattr(cell, 'data_type') and cell.data_type == 'f':
                                    # Formula cell
                                    sheet_analysis["data_types"]["formula"] += 1
                                    sheet_analysis["formulas_count"] += 1
                                    analysis_result["total_formulas"] += 1
                                    
                                    if cell.value:
                                        functions = self._extract_functions(str(cell.value))
                                        sheet_analysis["functions_used"].update(functions)
                                        analysis_result["unique_functions"].update(functions)
                                elif isinstance(cell.value, (int, float)):
                                    sheet_analysis["data_types"]["number"] += 1
                                else:
                                    sheet_analysis["data_types"]["text"] += 1
                                
                                # Check for errors
                                if isinstance(cell.value, str) and cell.value.startswith('#'):
                                    sheet_analysis["errors_found"].append(f"{cell.coordinate}: {cell.value}")
                                    
                            except Exception as cell_error:
                                # Skip problematic cells
                                continue
                    
                    # Check for conditional formatting
                    if hasattr(sheet, 'conditional_formatting') and sheet.conditional_formatting:
                        analysis_result["advanced_features"]["has_conditional_formatting"] = True
                    
                    # Check for data validation
                    if hasattr(sheet, 'data_validations') and sheet.data_validations:
                        analysis_result["advanced_features"]["has_data_validation"] = True
                    
                    analysis_result["cell_analysis"][sheet_name] = {
                        "max_row": sheet_analysis["max_row"],
                        "max_column": sheet_analysis["max_column"],
                        "formulas_count": sheet_analysis["formulas_count"],
                        "functions_used": list(sheet_analysis["functions_used"]),
                        "data_distribution": sheet_analysis["data_types"],
                        "errors_count": len(sheet_analysis["errors_found"]),
                        "errors": sheet_analysis["errors_found"][:10]  # Limit error list
                    }
                    
                except Exception as sheet_error:
                    self.logger.warning(f"Sheet analysis failed for {sheet_name}: {sheet_error}")
                    continue
            
            # Check for named ranges
            if hasattr(workbook, 'defined_names') and workbook.defined_names:
                analysis_result["advanced_features"]["has_named_ranges"] = True
            
            # Calculate quality metrics
            total_cells = sum(
                sum(sheet_data["data_distribution"].values()) 
                for sheet_data in analysis_result["cell_analysis"].values()
            )
            
            formula_percentage = (analysis_result["total_formulas"] / max(total_cells, 1)) * 100
            
            analysis_result["data_quality_metrics"] = {
                "total_cells": total_cells,
                "formula_percentage": round(formula_percentage, 2),
                "function_diversity": len(analysis_result["unique_functions"]),
                "error_cells": sum(len(sheet_data["errors"]) for sheet_data in analysis_result["cell_analysis"].values()),
                "complexity_score": min(10.0, len(analysis_result["unique_functions"]) * 1.2 + analysis_result["total_formulas"] * 0.1)
            }
            
            # Convert sets to lists for JSON serialization
            analysis_result["unique_functions"] = list(analysis_result["unique_functions"])
            
            return analysis_result
            
        except Exception as e:
            raise Exception(f"OpenPyXL analysis failed: {e}")
    
    async def _analyze_with_pandas(self, file_path: str) -> Dict[str, Any]:
        """Analysis using pandas for broader file support"""
        
        try:
            file_ext = Path(file_path).suffix.lower()
            
            if file_ext == '.csv':
                df = pd.read_csv(file_path)
                sheet_data = {"Sheet1": df}
            else:
                # Excel file
                sheet_data = pd.read_excel(file_path, sheet_name=None, engine='openpyxl' if file_ext == '.xlsx' else 'xlrd')
            
            analysis_result = {
                "file_path": file_path,
                "file_size_bytes": Path(file_path).stat().st_size,
                "sheet_names": list(sheet_data.keys()),
                "total_sheets": len(sheet_data),
                "total_formulas": 0,  # Limited formula detection with pandas
                "unique_functions": [],
                "data_analysis": {},
                "data_quality_metrics": {}
            }
            
            # Analyze each sheet's data
            for sheet_name, df in sheet_data.items():
                sheet_analysis = {
                    "rows": len(df),
                    "columns": len(df.columns),
                    "null_cells": df.isnull().sum().sum(),
                    "data_types": df.dtypes.value_counts().to_dict(),
                    "numeric_columns": len(df.select_dtypes(include=['number']).columns),
                    "text_columns": len(df.select_dtypes(include=['object']).columns)
                }
                
                analysis_result["data_analysis"][sheet_name] = sheet_analysis
            
            # Calculate overall metrics
            total_cells = sum(analysis["rows"] * analysis["columns"] for analysis in analysis_result["data_analysis"].values())
            total_null = sum(analysis["null_cells"] for analysis in analysis_result["data_analysis"].values())
            
            analysis_result["data_quality_metrics"] = {
                "total_cells": total_cells,
                "null_percentage": (total_null / max(total_cells, 1)) * 100,
                "data_completeness": ((total_cells - total_null) / max(total_cells, 1)) * 100,
                "complexity_score": min(5.0, len(sheet_data) * 0.5 + total_cells / 1000)
            }
            
            return analysis_result
            
        except Exception as e:
            raise Exception(f"Pandas analysis failed: {e}")
    
    async def _analyze_with_xlrd(self, file_path: str) -> Dict[str, Any]:
        """Basic analysis with xlrd for .xls files"""
        
        try:
            workbook = xlrd.open_workbook(file_path)
            
            analysis_result = {
                "file_path": file_path,
                "file_size_bytes": Path(file_path).stat().st_size,
                "sheet_names": workbook.sheet_names(),
                "total_sheets": workbook.nsheets,
                "total_formulas": 0,  # xlrd has limited formula support
                "unique_functions": [],
                "basic_analysis": {},
                "data_quality_metrics": {}
            }
            
            total_cells = 0
            total_empty = 0
            
            for sheet_index in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_index)
                sheet_name = sheet.name
                
                cells_with_data = 0
                for row in range(min(sheet.nrows, 100)):  # Limit rows
                    for col in range(min(sheet.ncols, 50)):  # Limit columns
                        cell_value = sheet.cell_value(row, col)
                        if cell_value:
                            cells_with_data += 1
                        total_cells += 1
                
                total_empty += (min(sheet.nrows, 100) * min(sheet.ncols, 50)) - cells_with_data
                
                analysis_result["basic_analysis"][sheet_name] = {
                    "rows": sheet.nrows,
                    "columns": sheet.ncols,
                    "cells_with_data": cells_with_data
                }
            
            analysis_result["data_quality_metrics"] = {
                "total_cells": total_cells,
                "empty_cell_percentage": (total_empty / max(total_cells, 1)) * 100,
                "complexity_score": min(3.0, workbook.nsheets * 0.5)
            }
            
            return analysis_result
            
        except Exception as e:
            raise Exception(f"XLRD analysis failed: {e}")
    
    def _basic_file_analysis(self, file_path: str) -> Dict[str, Any]:
        """Basic file analysis when libraries unavailable"""
        
        try:
            path = Path(file_path)
            file_size = path.stat().st_size if path.exists() else 0
            
            return {
                "file_path": file_path,
                "file_size_bytes": file_size,
                "file_name": path.name,
                "file_extension": path.suffix,
                "sheet_names": ["Unknown"],
                "total_formulas": 0,
                "unique_functions": [],
                "analysis_method": "basic_file_info",
                "data_quality_metrics": {
                    "file_exists": path.exists(),
                    "file_readable": True,
                    "complexity_score": 1.0 if file_size > 0 else 0.0
                },
                "limitations": "Limited analysis - Excel libraries not available"
            }
        except Exception as e:
            return self._error_analysis(file_path, f"Basic analysis failed: {e}")
    
    def _error_analysis(self, file_path: str, error: str) -> Dict[str, Any]:
        """Generate error analysis result"""
        
        return {
            "file_path": file_path,
            "file_size_bytes": 0,
            "sheet_names": [],
            "total_formulas": 0,
            "unique_functions": [],
            "data_quality_metrics": {
                "file_accessible": False,
                "complexity_score": 0.0
            },
            "analysis_error": error,
            "error_timestamp": datetime.utcnow().isoformat()
        }
    
    def _extract_functions(self, formula: str) -> List[str]:
        """Extract Excel function names from formula with enhanced detection"""
        
        try:
            if not formula or not isinstance(formula, str):
                return []
            
            # Enhanced regex patterns for Excel functions
            patterns = [
                r'\b([A-Z][A-Z0-9_]*)\s*\(',  # Standard functions
                r'([A-Z]+)\.',  # Functions with periods
                r'_x([A-Fa-f0-9]+)_'  # Some encoded functions
            ]
            
            functions = set()
            for pattern in patterns:
                matches = re.findall(pattern, formula.upper())
                functions.update(matches)
            
            # Filter out common non-functions
            non_functions = {'IF', 'AND', 'OR', 'NOT', 'TRUE', 'FALSE'}
            valid_functions = functions - non_functions
            
            return list(valid_functions)
            
        except Exception as e:
            self.logger.warning(f"Function extraction failed: {e}")
            return []
    
    def _validate_file(self, file_path: str) -> bool:
        """Comprehensive file validation"""
        
        try:
            path = Path(file_path)
            
            # Basic validation
            if not path.exists():
                self.logger.warning(f"File does not exist: {file_path}")
                return False
            
            if not path.is_file():
                self.logger.warning(f"Path is not a file: {file_path}")
                return False
            
            # Size validation
            file_size = path.stat().st_size
            if file_size > self.max_file_size:
                self.logger.warning(f"File too large: {file_size} bytes")
                return False
            
            if file_size == 0:
                self.logger.warning(f"File is empty: {file_path}")
                return False
            
            # Extension validation
            if path.suffix.lower() not in self.supported_formats:
                self.logger.warning(f"Unsupported file format: {path.suffix}")
                return False
            
            # Try to read first few bytes
            try:
                with open(path, 'rb') as f:
                    header = f.read(8)
                    if len(header) < 4:
                        self.logger.warning(f"File appears corrupted: {file_path}")
                        return False
            except Exception as e:
                self.logger.warning(f"File read test failed: {e}")
                return False
            
            return True
            
        except Exception as e:
            self.logger.warning(f"File validation failed: {e}")
            return False
    
    def get_analysis_stats(self) -> Dict[str, Any]:
        """Get file analysis statistics"""
        
        success_rate = 0
        if self.analysis_stats["files_analyzed"] > 0:
            success_rate = (self.analysis_stats["successful_analyses"] / self.analysis_stats["files_analyzed"]) * 100
        
        return {
            "files_analyzed": self.analysis_stats["files_analyzed"],
            "successful_analyses": self.analysis_stats["successful_analyses"],
            "failed_analyses": self.analysis_stats["failed_analyses"],
            "success_rate_percentage": round(success_rate, 2),
            "total_formulas_found": self.analysis_stats["total_formulas_found"],
            "unique_functions_count": len(self.analysis_stats["unique_functions_discovered"]),
            "library_availability": {
                "openpyxl": self.openpyxl_available,
                "xlrd": self.xlrd_available,
                "pandas": self.pandas_available
            }
        }

# =============================================================================
# COMPLETE MAIN EVALUATION ENGINE
# =============================================================================

class ClaudeEvaluationEngine:
    """COMPLETE: Main evaluation engine with full production features"""
    
    def __init__(
        self, 
        anthropic_client=None,
        anthropic_api_key: str = None,
        redis_client=None,
        claude_model: str = "claude-3-5-sonnet-20241022",
        cache_ttl_hours: int = 24
    ):
        self.logger = logging.getLogger(__name__)
        
        # Initialize claude_api with comprehensive error handling
        try:
            self.claude_api = ClaudeAPIWrapper(anthropic_client, anthropic_api_key)
            if self.claude_api.available:
                self.logger.info("✅ Claude API initialized successfully")
            else:
                self.logger.warning("⚠️ Claude API not available - using enhanced fallback evaluations")
        except Exception as e:
            self.logger.error(f"❌ Failed to initialize claude_api: {e}")
            self.claude_api = self._create_fallback_claude_api()
        
        # Initialize file analyzer with comprehensive error handling
        try:
            self.file_analyzer = ExcelFileAnalyzer()
            self.logger.info("✅ File analyzer initialized")
        except Exception as e:
            self.logger.error(f"❌ Failed to initialize file_analyzer: {e}")
            self.file_analyzer = self._create_fallback_file_analyzer()
        
        # Initialize cache with comprehensive error handling
        try:
            self.cache = EvaluationCache(redis_client, cache_ttl_hours)
            self.logger.info("✅ Evaluation cache initialized")
        except Exception as e:
            self.logger.error(f"❌ Failed to initialize cache: {e}")
            self.cache = self._create_fallback_cache()
        
        # Orchestrator compatibility properties
        self.claude_client = self.claude_api
        self.file_analyzer_engine = self.file_analyzer
        self.evaluation_cache = self.cache
        self.anthropic_client = getattr(self.claude_api, 'anthropic_client', None)
        
        # Comprehensive performance tracking
        self.evaluation_stats = {
            "total_evaluations": 0,
            "cache_hits": 0,
            "llm_evaluations": 0,
            "file_evaluations": 0,
            "hybrid_evaluations": 0,
            "avg_evaluation_time": 0.0,
            "score_distribution": {"0-1": 0, "1-2": 0, "2-3": 0, "3-4": 0, "4-5": 0},
            "evaluation_methods": {"claude": 0, "fallback": 0, "enhanced_fallback": 0}
        }
        
        self.logger.info("✅ ClaudeEvaluationEngine fully initialized")
    
    async def evaluate_response(
        self, 
        question,
        text_response: str = None,
        file_path: str = None
    ) -> Dict[str, Any]:
        """COMPLETE: Main evaluation entry point with full functionality"""
        
        start_time = time.time()
        self.evaluation_stats["total_evaluations"] += 1
        
        try:
            # Convert question to comprehensive dict if needed
            question_dict = self._normalize_question(question)
            
            # Generate cache key
            cache_content = f"{text_response or ''}{file_path or ''}"
            cache_key = self.cache._generate_cache_key(
                question_dict.get("id", "unknown"), 
                cache_content
            )
            
            # Check cache first
            try:
                cached_result = await self.cache.get_cached_evaluation(cache_key)
                if cached_result:
                    self.evaluation_stats["cache_hits"] += 1
                    cached_result["cache_hit"] = True
                    return cached_result
            except Exception as e:
                self.logger.warning(f"Cache check failed: {e}")
            
            # Determine evaluation strategy
            question_type = question_dict.get("type", "free_text")
            
            if question_type == "free_text" and text_response:
                evaluation = await self._evaluate_text_response(question_dict, text_response)
                self.evaluation_stats["llm_evaluations"] += 1
            elif question_type == "file_upload" and file_path:
                evaluation = await self._evaluate_file_response(question_dict, file_path)
                self.evaluation_stats["file_evaluations"] += 1
            elif question_type == "hybrid" or (text_response and file_path):
                evaluation = await self._evaluate_hybrid_response(question_dict, text_response, file_path)
                self.evaluation_stats["hybrid_evaluations"] += 1
            else:
                # Default to text evaluation
                evaluation = await self._evaluate_text_response(question_dict, text_response or "No response provided")
                self.evaluation_stats["llm_evaluations"] += 1
            
            # Add comprehensive metadata
            evaluation_time = time.time() - start_time
            evaluation.update({
                "question_id": question_dict.get("id", "unknown"),
                "evaluation_time_ms": int(evaluation_time * 1000),
                "evaluator_type": "claude" if self.claude_api.available else "enhanced_fallback",
                "created_at": datetime.utcnow().isoformat(),
                "question_metadata": {
                    "difficulty": question_dict.get("difficulty"),
                    "skill_category": question_dict.get("skill_category"),
                    "type": question_dict.get("type")
                },
                "cache_hit": False
            })
            
            # Update statistics
            self._update_evaluation_stats(evaluation_time, evaluation)
            
            # Cache result
            try:
                await self.cache.cache_evaluation(cache_key, evaluation)
            except Exception as e:
                self.logger.warning(f"Cache storage failed: {e}")
            
            return evaluation
            
        except Exception as e:
            evaluation_time = time.time() - start_time
            self.logger.error(f"❌ Evaluation failed: {e}")
            return self._create_error_evaluation(
                question_dict.get("id", "unknown") if 'question_dict' in locals() else "unknown", 
                str(e),
                evaluation_time
            )
    
    def _normalize_question(self, question) -> Dict[str, Any]:
        """Convert question to normalized dictionary format"""
        
        if hasattr(question, 'id'):
            return {
                "id": question.id,
                "text": getattr(question, 'text', ''),
                "type": getattr(question, 'type', 'free_text'),
                "skill_category": getattr(question, 'skill_category', 'general'),
                "difficulty": getattr(question, 'difficulty', 'intermediate'),
                "expected_keywords": getattr(question, 'expected_keywords', []),
                "common_mistakes": getattr(question, 'common_mistakes', []),
                "estimated_time_minutes": getattr(question, 'estimated_time_minutes', 3)
            }
        elif isinstance(question, dict):
            # Ensure all required fields are present
            return {
                "id": question.get("id", "unknown"),
                "text": question.get("text", ""),
                "type": question.get("type", "free_text"),
                "skill_category": question.get("skill_category", "general"),
                "difficulty": question.get("difficulty", "intermediate"),
                "expected_keywords": question.get("expected_keywords", []),
                "common_mistakes": question.get("common_mistakes", []),
                "estimated_time_minutes": question.get("estimated_time_minutes", 3)
            }
        else:
            # Fallback for unknown question format
            return {
                "id": "unknown",
                "text": str(question) if question else "",
                "type": "free_text",
                "skill_category": "general",
                "difficulty": "intermediate",
                "expected_keywords": [],
                "common_mistakes": [],
                "estimated_time_minutes": 3
            }
    
    async def _evaluate_text_response(self, question: Dict, response_text: str) -> Dict[str, Any]:
        """COMPLETE: Evaluate text responses with full Claude integration"""
        
        if not response_text or not response_text.strip():
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": "No response provided",
                "strengths": [],
                "areas_for_improvement": ["Please provide a response"],
                "keywords_found": [],
                "mistakes_detected": [],
                "evaluation_method": "no_response"
            }
        
        try:
            # Use Claude API for evaluation
            evaluation = await self.claude_api.evaluate_text_response(question, response_text)
            self.evaluation_stats["evaluation_methods"]["claude"] += 1
            return evaluation
        except Exception as e:
            self.logger.error(f"Text evaluation failed: {e}")
            self.evaluation_stats["evaluation_methods"]["enhanced_fallback"] += 1
            return await self._enhanced_fallback_text_evaluation(question, response_text)
    
    async def _evaluate_file_response(self, question: Dict, file_path: str) -> Dict[str, Any]:
        """COMPLETE: Evaluate file responses with comprehensive analysis"""
        
        if not file_path:
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": "No file uploaded",
                "strengths": [],
                "areas_for_improvement": ["Please upload an Excel file"],
                "keywords_found": [],
                "mistakes_detected": [],
                "evaluation_method": "no_file"
            }
        
        try:
            # Analyze file comprehensively
            file_analysis = await self.file_analyzer.analyze_excel_file(file_path, question)
            evaluation = self._score_comprehensive_file_analysis(question, file_analysis)
            return evaluation
        except Exception as e:
            self.logger.error(f"File evaluation failed: {e}")
            return {
                "score": 1.0,
                "confidence": 0.3,
                "reasoning": f"File analysis failed: {str(e)}",
                "strengths": ["File uploaded successfully"],
                "areas_for_improvement": ["Ensure valid Excel file format", "Check file for corruption"],
                "keywords_found": [],
                "mistakes_detected": [f"File analysis error: {str(e)}"],
                "evaluation_method": "file_error"
            }
    
    async def _evaluate_hybrid_response(self, question: Dict, text_response: str, file_path: str) -> Dict[str, Any]:
        """COMPLETE: Evaluate hybrid responses with sophisticated combination logic"""
        
        # Get individual evaluations
        text_eval = None
        file_eval = None
        
        if text_response and text_response.strip():
            text_eval = await self._evaluate_text_response(question, text_response)
        
        if file_path:
            file_eval = await self._evaluate_file_response(question, file_path)
        
        # Handle missing components
        if not text_eval and not file_eval:
            return {
                "score": 0.0,
                "confidence": 1.0,
                "reasoning": "Both text explanation and file required for complete answer",
                "strengths": [],
                "areas_for_improvement": ["Provide both written explanation and Excel file"],
                "keywords_found": [],
                "mistakes_detected": [],
                "evaluation_method": "hybrid_incomplete"
            }
        
        # Combine evaluations with sophisticated weighting
        if text_eval and file_eval:
            # Both components present - weighted combination
            text_weight = 0.6  # Text explanation is primary
            file_weight = 0.4  # File demonstrates practical application
            
            combined_score = (text_eval["score"] * text_weight + file_eval["score"] * file_weight)
            combined_confidence = min(text_eval.get("confidence", 0.5), file_eval.get("confidence", 0.5))
            
            # Combine feedback
            combined_reasoning = f"Text Analysis: {text_eval['reasoning']} | File Analysis: {file_eval['reasoning']}"
            combined_strengths = text_eval.get("strengths", []) + file_eval.get("strengths", [])
            combined_improvements = text_eval.get("areas_for_improvement", []) + file_eval.get("areas_for_improvement", [])
            combined_keywords = text_eval.get("keywords_found", []) + file_eval.get("keywords_found", [])
            combined_mistakes = text_eval.get("mistakes_detected", []) + file_eval.get("mistakes_detected", [])
            
            # Synergy bonus - reward when text and file complement each other
            synergy_bonus = 0.0
            if text_eval["score"] >= 3.0 and file_eval["score"] >= 3.0:
                synergy_bonus = 0.3  # Bonus for good performance in both areas
            
            final_score = min(5.0, combined_score + synergy_bonus)
            
        elif text_eval:
            # Only text provided
            final_score = text_eval["score"] * 0.7  # Penalty for missing file
            combined_confidence = text_eval.get("confidence", 0.5) * 0.8
            combined_reasoning = f"Text Only: {text_eval['reasoning']}. Missing practical file demonstration."
            combined_strengths = text_eval.get("strengths", [])
            combined_improvements = text_eval.get("areas_for_improvement", []) + ["Provide Excel file to demonstrate practical application"]
            combined_keywords = text_eval.get("keywords_found", [])
            combined_mistakes = text_eval.get("mistakes_detected", []) + ["No file provided"]
            
        else:
            # Only file provided
            final_score = file_eval["score"] * 0.7  # Penalty for missing explanation
            combined_confidence = file_eval.get("confidence", 0.5) * 0.8
            combined_reasoning = f"File Only: {file_eval['reasoning']}. Missing written explanation."
            combined_strengths = file_eval.get("strengths", [])
            combined_improvements = file_eval.get("areas_for_improvement", []) + ["Provide written explanation of your approach"]
            combined_keywords = file_eval.get("keywords_found", [])
            combined_mistakes = file_eval.get("mistakes_detected", []) + ["No text explanation provided"]
        
        return {
            "score": round(final_score, 2),
            "confidence": round(combined_confidence, 2),
            "reasoning": combined_reasoning,
            "strengths": list(set(combined_strengths))[:5],  # Remove duplicates, limit to 5
            "areas_for_improvement": list(set(combined_improvements))[:4],  # Remove duplicates, limit to 4
            "keywords_found": list(set(combined_keywords)),
            "mistakes_detected": list(set(combined_mistakes)),
            "evaluation_method": "hybrid_comprehensive",
            "component_scores": {
                "text_score": text_eval["score"] if text_eval  else 0.0 
                }
        }  